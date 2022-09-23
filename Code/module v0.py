# -*- coding: utf-8 -*-
"""
Created on Fri Sep  9 13:58:26 2022

@author: a7086
"""

import scipy
import math
import numpy as np
import pandas as pd
import scipy.stats
from scipy.optimize import minimize 
import numpy.linalg as lin
from scipy.optimize import fsolve


def Data_Generate(cros_type,mean_true,sample_size,data_type,cor_par,eta0):
    #mean_true=true_mean
    #sample_size=100000
    #cor_par=gamma_param
    #sample_size=int(sample_size/2)
    ''' 假設樣本之間為獨立生成的Poisson分配，求其MLE、I matrix、V matrix
    Input：
        -cros_type:交叉設計方式
        -mean_true:用參數真值算出來的樣本平均數，用以生成獨立卜瓦松樣本y11、y21、y12、y22
        -sample_size：yij的樣本數(總樣本數為sample_sizs*4)
        -data_type:ind or cor
        
    Output：data
    '''
    if len(cros_type)==4:
        col_res=['Yi11', 'Yi12','Yi21', 'Yi22']
        col_mu=['Mu11', 'Mu12','Mu21', 'Mu22']
    elif len(cros_type)==6:
        col_res=['Yi11', 'Yi12','Yi13','Yi21','Yi22', 'Yi23']
        col_mu=['Mu11', 'Mu12','Mu13','Mu21','Mu22', 'Mu23']
    elif len(cros_type)==9:
        col_res=['Yi11', 'Yi12','Yi13','Yi21','Yi22', 'Yi23','Yi31','Yi32', 'Yi33']
        col_mu=['Mu11', 'Mu12','Mu13','Mu21','Mu22', 'Mu23','Mu31','Mu32', 'Mu33']
    
    if data_type == 'ind': 
        data=pd.DataFrame(np.array([np.random.poisson(lam=p, size=sample_size) for p in mean_true]).T.tolist(),columns =col_res)
    else:
        '''
        #np.random.seed(980716)
        data=Corr_Poisson(mean_true=mean_true,sample_size=sample_size, cor_par=cor_par)
        #data.corr()
        '''
        
        #np.random.seed(980716)
        Mu_cor = pd.DataFrame(columns = col_mu)
        for i in range(0,len(mean_true),2):            
            nu=np.random.gamma(1/cor_par,cor_par,sample_size)
            #print(i,nu)
            for (ix,param) in enumerate(mean_true[i:i + 2]):
                Mu_cor[Mu_cor.columns[ix+i]]=pd.DataFrame(np.multiply(np.array(param).repeat(sample_size), nu).T)
                #print(i,ix,param)

        data=pd.DataFrame(columns = col_res)
        for (idx,mu) in enumerate(Mu_cor.columns):
            data[data.columns[idx]]=pd.DataFrame([np.random.poisson(p) for p in Mu_cor[mu]])
            
    return data
 

def Mean_Poisson(cros_type,params_value):
    '''
    用途:計算平均值
    1.cors_type:交叉設計
    2.params_value:若帶入的是真值則為真實平均值，若為MLE則為估計量
    
    '''
    #找出交叉設計中B藥與C藥的組別與時間點
    loc_B=np.array([[i,1] for i, w in enumerate(cros_type) if 'B' in w])
    loc_C=np.array([[i,2] for i, w in enumerate(cros_type) if 'C' in w])
    #只有B藥
    if len(loc_C) == 0:
        #2X2 e.g. ABBA
        if len(cros_type)==4:
            tao,eta,gamma,delta=params_value
            params = np.array([ [tao], [eta],  [gamma], [delta]])
            covariate=np.array([[1,0,0,0],[1,0,1,0],[1,0,0,1],[1,0,1,1]])
            covariate[tuple(loc_B.T)] = 1
        #2X3 e.g. ABBBAA
        if len(cros_type)==6:
            tao,eta,gamma1,gamma2,delta=params_value
            params = np.array([ [tao], [eta], [gamma1], [gamma2], [delta]])
            covariate=np.array([[1,0,0,0,0],[1,0,1,0,0],[1,0,0,1,0],[1,0,0,0,1],[1,0,1,0,1],[1,0,0,1,1]])
            covariate[tuple(loc_B.T)] = 1
        #3X3 e.g.AABABABAA
        if len(cros_type)==9:
            tao,eta,gamma1,gamma2,delta1,delta2=params_value
            params = np.array([ [tao], [eta], [gamma1], [gamma2], [delta1], [delta2]])
            covariate=np.array([[1,0,0,0,0,0],[1,0,1,0,0,0],[1,0,0,1,0,0],[1,0,0,0,1,0],[1,0,1,0,1,0],[1,0,0,1,1,0],[1,0,0,0,0,1],[1,0,1,0,0,1],[1,0,0,1,0,1]])
            covariate[tuple(loc_B.T)] = 1
    #3X3 有B藥與C藥    
    elif (len(loc_C) != 0) and (len(cros_type)==9):
        tao,eta1,eta2,gamma1,gamma2,delta1,delta2=params_value
        params = np.array([ [tao], [eta1], [eta2], [gamma1], [gamma2], [delta1], [delta2]])
        covariate=np.array([[1,0,0,0,0,0,0],[1,0,0,1,0,0,0],[1,0,0,0,1,0,0],[1,0,0,0,0,1,0],[1,0,0,1,0,1,0],[1,0,0,0,1,1,0],[1,0,0,0,0,0,1],[1,0,0,1,0,0,1],[1,0,0,0,1,0,1]])
        covariate[tuple(loc_B.T)] = 1
        covariate[tuple(loc_C.T)] = 1
    return np.exp(np.array([ xijk  for xijk in covariate]).dot(params)).reshape(1,len(cros_type)).tolist()[0]
  

def MLE(cros_type,data,initial_guess):
    '''
    用途：計算矩陣V(Poisson模型推導結果)
    變數解釋：
    1.cros_type:序列設計方式 e.g.'ABBA'
    2.data:生成的資料
    3.initial_guess
    initial_guess = np.array([0.841, 0.318, 0.11, 0.223, 0.202, 0.262])
    initial_guess = np.array([0.95, 0.372, 0.354, 0.236, 0.226, 0.206, 0.264])
    '''
    #3X3 pi=(seq_size)/(seq_size*3)
    #3X2、2X2 pi=(seq_size)/(seq_size*2)
    if cros_type=='ABBA':
        tao_hat=np.log(np.mean(data['Yi11']))
        eta_hat=0.5*(np.log(np.mean(data['Yi21']))+np.log(np.mean(data['Yi12']))-np.log(np.mean(data['Yi11']))-np.log(np.mean(data['Yi22'])))
        gamma_hat=0.5*(np.log(np.mean(data['Yi12']))+np.log(np.mean(data['Yi22']))-np.log(np.mean(data['Yi11']))-np.log(np.mean(data['Yi21'])))
        delta_hat=0.5*(np.log(np.mean(data['Yi21']))+np.log(np.mean(data['Yi22']))-np.log(np.mean(data['Yi11']))-np.log(np.mean(data['Yi12'])))
        
        # MLE
        estimate= pd.DataFrame({'tao_hat': tao_hat, 
                                'eta_hat': eta_hat, 
                                'gamma_hat':gamma_hat, 
                                'delta_hat': delta_hat},index=[0])
    elif cros_type=='ABBBAA':       
        tao_hat = np.log(np.mean(data['Yi11']))
        eta_hat =0.25*(np.log(np.mean(data['Yi12']))+np.log(np.mean(data['Yi13']))+2*np.log(np.mean(data['Yi21']))-np.log(np.mean(data['Yi22']))-np.log(np.mean(data['Yi23']))-2*np.log(np.mean(data['Yi11'])))
        gamma1_hat = 0.5*(np.log(np.mean(data['Yi12']))+np.log(np.mean(data['Yi22']))-np.log(np.mean(data['Yi21']))-np.log(np.mean(data['Yi11'])))
        gamma2_hat = 0.5*(np.log(np.mean(data['Yi13']))+np.log(np.mean(data['Yi23']))-np.log(np.mean(data['Yi21']))-np.log(np.mean(data['Yi11'])))
        delta_hat = 0.5*(np.log(np.mean(data['Yi21']))-np.log(np.mean(data['Yi11'])))-0.25*(np.log(np.mean(data['Yi12']))+np.log(np.mean(data['Yi13']))-np.log(np.mean(data['Yi22']))-np.log(np.mean(data['Yi23'])))
        
        # MLE
        estimate= pd.DataFrame({'tao': tao_hat, 
                                'eta_hat': eta_hat, 
                                'gamma1_hat':gamma1_hat, 
                                'gamma2_hat':gamma2_hat,
                                'delta_hat': delta_hat},index=[0])
    elif cros_type=='AABABABAA':
        def L_332(params):
            tao, eta, gamma1, gamma2, delta1, delta2=params
            f1,f2,f3=0,0,0
            for i in range(len(data)):
                f1=f1+tao*data.at[i,'Yi11']+(tao+gamma1)*data.at[i,'Yi21']+(tao+eta+gamma2)*data.at[i,'Yi31']-np.exp(tao)-np.exp(tao+gamma1)-np.exp(tao+eta+gamma2)
                f2=f2+(tao+delta1)*data.at[i,'Yi12']+(tao+eta+gamma1+delta1)*data.at[i,'Yi22']+(tao+gamma2+delta1)*data.at[i,'Yi32']-np.exp(tao+delta1)-np.exp(tao+eta+gamma1+delta1)-np.exp(tao+gamma2+delta1)
                f3=f3+(tao+eta+delta2)*data.at[i,'Yi13']+(tao+gamma1+delta2)*data.at[i,'Yi23']+(tao+gamma2+delta2)*data.at[i,'Yi33']-np.exp(tao+eta+delta2)-np.exp(tao+gamma1+delta2)-np.exp(tao+gamma2+delta2)
            return f1+f2+f3
        solve = scipy.optimize.minimize (lambda params: -L_332(params), initial_guess)
        MLE_opt=solve.x
        estimate= pd.DataFrame({'tao_hat': MLE_opt[0], 
                                'eta_hat': MLE_opt[1], 
                                'gamma1_hat': MLE_opt[2], 
                                'gamma2_hat': MLE_opt[3], 
                                'delta1_hat': MLE_opt[4],
                                'delta2_hat': MLE_opt[5],
                                },index=[0])
    elif cros_type=='ABCBCACAB':
        def L_3331(params):
            tao, eta1, eta2, gamma1, gamma2, delta1, delta2=params
            f1,f2,f3=0,0,0
            for i in range(len(data)):
                f1=f1+tao*data.at[i,'Yi11']+(tao+eta1+gamma1)*data.at[i,'Yi21']+(tao+eta2+gamma2)*data.at[i,'Yi31']-np.exp(tao)-np.exp(tao+eta1+gamma1)-np.exp(tao+eta2+gamma2)
                f2=f2+(tao+eta1+delta1)*data.at[i,'Yi12']+(tao+eta2+gamma1+delta1)*data.at[i,'Yi22']+(tao+gamma2+delta1)*data.at[i,'Yi32']-np.exp(tao+eta1+delta1)-np.exp(tao+eta2+gamma1+delta1)-np.exp(tao+gamma2+delta1)
                f3=f3+(tao+eta2+delta2)*data.at[i,'Yi13']+(tao+gamma1+delta2)*data.at[i,'Yi23']+(tao+eta1+gamma2+delta2)*data.at[i,'Yi33']-np.exp(tao+eta2+delta2)-np.exp(tao+gamma1+delta2)-np.exp(tao+eta1+gamma2+delta2)
            return f1+f2+f3
        solve = scipy.optimize.minimize (lambda params: -L_3331(params), initial_guess)
        MLE_opt=solve.x
        estimate= pd.DataFrame({'tao_hat': MLE_opt[0], 
                                'eta1_hat': MLE_opt[1], 
                                'eta2_hat': MLE_opt[2], 
                                'gamma1_hat': MLE_opt[3], 
                                'gamma2_hat': MLE_opt[4],
                                'delta1_hat': MLE_opt[5],
                                'delta2_hat': MLE_opt[6]},index=[0])
    elif cros_type=='BACACBBCA':        
        def L_3332(params):
            tao, eta1, eta2, gamma1, gamma2, delta1, delta2=params
            f1,f2,f3=0,0,0
            for i in range(len(data)):
                f1=f1+(tao+eta1)*data.at[i,'Yi11']+(tao+gamma1)*data.at[i,'Yi21']+(tao+eta2+gamma2)*data.at[i,'Yi31']-np.exp(tao+eta1)-np.exp(tao+gamma1)-np.exp(tao+eta2+gamma2)
                f2=f2+(tao+delta1)*data.at[i,'Yi12']+(tao+eta2+gamma1+delta1)*data.at[i,'Yi22']+(tao+eta1+gamma2+delta1)*data.at[i,'Yi32']-np.exp(tao+delta1)-np.exp(tao+eta2+gamma1+delta1)-np.exp(tao+eta1+gamma2+delta1)
                f3=f3+(tao+eta1+delta2)*data.at[i,'Yi13']+(tao+eta2+gamma1+delta2)*data.at[i,'Yi23']+(tao+gamma2+delta2)*data.at[i,'Yi33']-np.exp(tao+eta1+delta2)-np.exp(tao+eta2+gamma1+delta2)-np.exp(tao+gamma2+delta2)
            return f1+f2+f3
        solve = scipy.optimize.minimize (lambda params: -L_3332(params), initial_guess)
        MLE_opt=solve.x
        estimate= pd.DataFrame({'tao_hat': MLE_opt[0], 
                                'eta1_hat': MLE_opt[1], 
                                'eta2_hat': MLE_opt[2], 
                                'gamma1_hat': MLE_opt[3], 
                                'gamma2_hat': MLE_opt[4],
                                'delta1_hat': MLE_opt[5],
                                'delta2_hat': MLE_opt[6]},index=[0])
    elif cros_type=='BBAACBCAC':
        def L_3333(params):
            tao, eta1, eta2, gamma1, gamma2, delta1, delta2=params
            f1,f2,f3=0,0,0
            for i in range(len(data)):
                f1=f1+(tao+eta1)*data.at[i,'Yi11']+(tao+eta1+gamma1)*data.at[i,'Yi21']+(tao+gamma2)*data.at[i,'Yi31']-np.exp(tao+eta1)-np.exp(tao+eta1+gamma1)-np.exp(tao+gamma2)
                f2=f2+(tao+delta1)*data.at[i,'Yi12']+(tao+eta2+gamma1+delta1)*data.at[i,'Yi22']+(tao+eta1+gamma2+delta1)*data.at[i,'Yi32']-np.exp(tao+delta1)-np.exp(tao+eta2+gamma1+delta1)-np.exp(tao+eta1+gamma2+delta1)
                f3=f3+(tao+eta2+delta2)*data.at[i,'Yi13']+(tao+gamma1+delta2)*data.at[i,'Yi23']+(tao+eta2+gamma2+delta2)*data.at[i,'Yi33']-np.exp(tao+eta2+delta2)-np.exp(tao+gamma1+delta2)-np.exp(tao+eta2+gamma2+delta2)
            return f1+f2+f3
        solve = scipy.optimize.minimize (lambda params: -L_3333(params), initial_guess)
        MLE_opt=solve.x
        estimate= pd.DataFrame({'tao_hat': MLE_opt[0], 
                                'eta1_hat': MLE_opt[1], 
                                'eta2_hat': MLE_opt[2], 
                                'gamma1_hat': MLE_opt[3], 
                                'gamma2_hat': MLE_opt[4],
                                'delta1_hat': MLE_opt[5],
                                'delta2_hat': MLE_opt[6]},index=[0])
    return estimate
    

def MatrixI(cros_type,Pi,tm):
    
    '''
    用途：計算矩陣I(Poisson模型推導結果)
    變數解釋：
    1.cros_type:序列設計方式 e.g.'ABBA'
    2.Pi:每個序列的人數/總人數
    3.tm:若放入的是真值則為I，若放入的是估計值則為I_hat
    '''
    #3X3 pi=(seq_size)/(seq_size*3)
    #3X2、2X2 pi=(seq_size)/(seq_size*2)
    if cros_type=='ABBA':
        #ABBA
        #true mean
        I = Pi*np.array([[sum(tm),tm[1]+tm[2],tm[1]+tm[3],tm[2]+tm[3]],
                         [tm[1]+tm[2],tm[1]+tm[2],tm[1],tm[2]],
                         [tm[1]+tm[3],tm[1],tm[1]+tm[3],tm[3]],
                         [tm[2]+tm[3],tm[2],tm[3],tm[2]+tm[3]]                 
                         ])
    elif cros_type=='ABBBAA':
        #ABBBAA
        #true mean
        I = Pi*np.array([[sum(tm),tm[1]+tm[2]+tm[3],tm[1]+tm[4],tm[2]+tm[5],sum(tm[3:6])],
                     [tm[1]+tm[2]+tm[3],tm[1]+tm[2]+tm[3],tm[1],tm[2],tm[3]],
                     [tm[1]+tm[4],tm[1],tm[1]+tm[4],0,tm[4]],
                     [tm[2]+tm[5],tm[2],0,tm[2]+tm[5],tm[5]],
                     [sum(tm[3:6]),tm[3],tm[4],tm[5],sum(tm[3:6])]                    
                    ])

    elif cros_type=='AABABABAA':
        #AABABABAA
        #true mean
        I = Pi*np.array([  [sum(tm), tm[2]+tm[4]+tm[6], tm[1]+tm[4]+tm[7], tm[2]+tm[5]+tm[8], sum(tm[3:6]), sum(tm[6:9])],
                         [tm[2]+tm[4]+tm[6], tm[2]+tm[4]+tm[6], tm[4], tm[2], tm[4], tm[6]],
                         [tm[1]+tm[4]+tm[7],tm[4],tm[1]+tm[4]+tm[7],0, tm[4],tm[7]],
                         [tm[2]+tm[5]+tm[8],tm[2],0,tm[2]+tm[5]+tm[8], tm[5], tm[8]],
                         [sum(tm[3:6]),tm[4],tm[4],tm[5],sum(tm[3:6]),0],
                         [sum(tm[6:9]),tm[6],tm[7],tm[8],0,sum(tm[6:9])]
                         ])
        
    elif cros_type=='ABCBCACAB':
        #ABCBCACAB
        #true mean
        I = Pi*np.array([[sum(tm),tm[1]+tm[3]+tm[8],tm[2]+tm[4]+tm[6],tm[1]+tm[4]+tm[7],tm[2]+tm[5]+tm[8],sum(tm[3:6]),sum(tm[6:9])],
                     [tm[1]+tm[3]+tm[8],tm[1]+tm[3]+tm[8],0, tm[1],tm[8],tm[3],tm[8]],
                     [tm[2]+tm[4]+tm[6],0 ,tm[2]+tm[4]+tm[6],tm[4], tm[2],tm[4],tm[6]],
                     [tm[1]+tm[4]+tm[7],tm[1],tm[4],tm[1]+tm[4]+tm[7],0, tm[4],tm[7]],
                     [tm[2]+tm[5]+tm[8],tm[8], tm[2],0,tm[2]+tm[5]+tm[8],tm[5],tm[5]],
                     [sum(tm[3:6]),tm[3],tm[4], tm[4],tm[5],sum(tm[3:6]),0],
                     [sum(tm[6:9]),tm[8],tm[6],tm[7],tm[5],0,sum(tm[6:9])]
                     ])
        
    elif cros_type=='BACACBBCA':
        #BACACBBCA
        #true mean
        I = Pi*np.array([[sum(tm),tm[0]+tm[5]+tm[6],tm[2]+tm[4]+tm[7],tm[1]+tm[4]+tm[7],tm[2]+tm[5]+tm[8],sum(tm[3:6]),sum(tm[6:9])],
                         [tm[0]+tm[5]+tm[6],tm[0]+tm[5]+tm[6] ,0 ,0 ,tm[5] ,tm[5] , tm[6]],
                         [tm[2]+tm[4]+tm[7],0 ,tm[2]+tm[4]+tm[7] ,tm[4]+tm[7] ,tm[2] , tm[4] , tm[7] ],
                         [tm[1]+tm[4]+tm[7],0 ,tm[4]+tm[7] ,tm[1]+tm[4]+tm[7] ,0 , tm[4] , tm[7]],
                         [tm[2]+tm[5]+tm[8], tm[5], tm[2] ,0 , tm[2]+tm[5]+tm[8], tm[5],tm[8] ],
                         [sum(tm[3:6]), tm[5],tm[4] ,tm[4] ,tm[5] , sum(tm[3:6]),0],
                         [sum(tm[6:9]), tm[6] , tm[7],tm[7] ,tm[8] ,0 ,sum(tm[6:9])]
                         ])
       
        
    elif cros_type=='BBAACBCAC':
        #BBAACBCAC
        #true mean
        I = Pi*np.array([[sum(tm),tm[0]+tm[1]+tm[5],tm[4]+tm[6]+tm[8],tm[1]+tm[4]+tm[7],tm[2]+tm[5]+tm[8],sum(tm[3:6]),sum(tm[6:9])],
                         [tm[0]+tm[1]+tm[5],tm[0]+tm[1]+tm[5],0,tm[1],tm[5],tm[5],0],
                         [tm[4]+tm[6]+tm[8],0,tm[4]+tm[6]+tm[8],tm[4],tm[8],tm[5],0],
                         [tm[1]+tm[4]+tm[7],tm[1],tm[4],tm[1]+tm[4]+tm[7],0,tm[4],tm[7]],
                         [tm[2]+tm[5]+tm[8],tm[5],tm[8],0,tm[2]+tm[5]+tm[8],tm[5],tm[8]],
                         [sum(tm[3:6]),tm[5],tm[5],tm[4],tm[5],sum(tm[3:6]),0],
                         [sum(tm[6:9]),0,0,tm[7],tm[8],0,sum(tm[6:9])]
                         ])
    return I


def MatrixV(cros_type,Pi,Cov,Var):
    '''
    用途：計算矩陣V(Poisson模型推導結果)
    變數解釋：
    1.cros_type:序列設計方式 e.g.'ABBA'
    2.Pi:每個序列的人數/總人數
    3.Cov:生成出的資料的Covariance
    4.Var:生成出的資料的Variance
    3&4需整理為整理格式如下
    Var=data.var(ddof=1)
    Cov=[ data.Yi11.cov(data.Yi21), data.Yi11.cov(data.Yi31),data.Yi21.cov(data.Yi31), 
          data.Yi12.cov(data.Yi22), data.Yi12.cov(data.Yi32),data.Yi22.cov(data.Yi32),
          data.Yi13.cov(data.Yi23), data.Yi13.cov(data.Yi33),data.Yi23.cov(data.Yi33)]
    '''
    #3X3 pi=(seq_size)/(seq_size*3)
    #3X2、2X2 pi=(seq_size)/(seq_size*2)
    if cros_type=='ABBA':
        V = Pi*np.array([
                    [sum(Var)+2*sum(Cov), Var[1]+Var[2]+sum(Cov), Var[1]+Var[3]+sum(Cov), sum(Var[2:4])+2*Cov[1]],
                    [Var[1]+Var[2]+sum(Cov),Var[1]+Var[2],Var[1]+Cov[1],Var[2]+Cov[1]],
                    [Var[1]+Var[3]+sum(Cov),Var[1]+Cov[1],Var[1]+Var[3],Var[3]+Cov[1]],
                    [sum(Var[2:4])+2*Cov[1],Var[2]+Cov[1],Var[3]+Cov[1],sum(Var[2:4])+2*Cov[1]]
                         ])
    elif cros_type=='ABBBAA':       
        V = pi*np.array([[sum(Var)+2*sum(Cov),Var[1]+Var[2]+Var[3]+sum(Cov)+Cov[2]-Cov[5],Var[1]+Var[4]+sum(Cov)-Cov[1]-Cov[4],Var[2]+Var[5]+sum(Cov)-Cov[0]-Cov[3],sum(Var[3:6])+2*sum(Cov[3:6])],
                     [Var[1]+Var[2]+Var[3]+sum(Cov)+Cov[2]-Cov[5],Var[1]+Var[2]+Var[3]+2*Cov[2],Var[1]+Cov[2]+Cov[3],Var[2]+Cov[2]+Cov[4],Var[3]+Cov[3]+Cov[4]],
                     [Var[1]+Var[4]+sum(Cov)-Cov[1]-Cov[4],Var[1]+Cov[2]+Cov[3],Var[1]+Var[4],Cov[2]+Cov[5],Var[4]+Cov[3]+Cov[5]],
                     [Var[2]+Var[5]+sum(Cov)-Cov[0]-Cov[3],Var[2]+Cov[2]+Cov[4],Cov[2]+Cov[5],Var[2]+Var[5],Var[5]+Cov[4]+Cov[5]],
                     [sum(Var[3:6])+2*sum(Cov[3:6]),Var[3]+Cov[3]+Cov[4],Var[4]+Cov[3]+Cov[5],Var[5]+Cov[4]+Cov[5],sum(Var[3:6])+2*sum(Cov[3:6])]
                    ])
    elif cros_type=='AABABABAA':
        V = pi*np.array([[sum(Var)+2*sum(Cov),Var[2]+Var[4]+Var[6]+sum(Cov)-Cov[0]-Cov[4]-Cov[8],Var[1]+Var[4]+Var[7]+sum(Cov)-Cov[1]-Cov[4]-Cov[7],Var[2]+Var[5]+Var[8]+sum(Cov)-Cov[0]-Cov[3]-Cov[6],sum(Var[3:6])+2*sum(Cov[3:6]),sum(Var[6:9])+2*sum(Cov[6:9])],
                     [Var[2]+Var[4]+Var[6]+sum(Cov)-Cov[0]-Cov[4]-Cov[8],Var[2]+Var[4]+Var[6],Var[4]+Cov[2]+Cov[6],Var[2]+Cov[5]+Cov[7],Var[4]+Cov[3]+Cov[5],Var[6]+Cov[6]+Cov[7]],
                     [Var[1]+Var[4]+Var[7]+sum(Cov)-Cov[1]-Cov[4]-Cov[7],Var[4]+Cov[2]+Cov[6],Var[1]+Var[4]+Var[7],Cov[2]+Cov[5]+Cov[8],Var[4]+Cov[3]+Cov[5],Var[7]+Cov[6]+Cov[8]],
                     [Var[2]+Var[5]+Var[8]+sum(Cov)-Cov[0]-Cov[3]-Cov[6],Var[2]+Cov[5]+Cov[7],Cov[2]+Cov[5]+Cov[8],Var[2]+Var[5]+Var[8],Var[5]+Cov[4]+Cov[5],Var[8]+Cov[7]+Cov[8]],
                     [sum(Var[3:6])+2*sum(Cov[3:6]),Var[4]+Cov[3]+Cov[5],Var[4]+Cov[3]+Cov[5],Var[5]+Cov[4]+Cov[5],sum(Var[3:6])+2*sum(Cov[3:6]),0],
                     [sum(Var[6:9])+2*sum(Cov[6:9]),Var[6]+Cov[6]+Cov[7],Var[7]+Cov[6]+Cov[8],Var[8]+Cov[7]+Cov[8],0,sum(Var[6:9])+2*sum(Cov[6:9])]
                    
                    ])
        
    elif cros_type=='ABCBCACAB':
        V = pi*np.array([[sum(Var)+2*sum(Cov),Var[1]+Var[3]+Var[8]+sum(Cov)-Cov[1]-Cov[5]-Cov[6],Var[2]+Var[4]+Var[6]+sum(Cov)-Cov[0]-Cov[4]-Cov[8],Var[1]+Var[4]+Var[7]+sum(Cov)-Cov[1]+Cov[2]-Cov[4]-Cov[7],Var[2]+Var[5]+Var[8]+sum(Cov)-Cov[0]-Cov[3]-Cov[6],sum(Var[3:6])+2*sum(Cov[3:6]),sum(Var[6:9])+2*sum(Cov[6:9])],
                     [Var[1]+Var[3]+Var[8]+sum(Cov)-Cov[1]-Cov[5]-Cov[6],Var[1]+Var[3]+Var[8],Cov[2]+Cov[3]+Cov[7],Var[2]+Cov[3]+Cov[8],Var[8]+Cov[2]+Cov[4],Var[3]+Cov[3]+Cov[4],Var[8]+Cov[7]+Cov[8]],
                     [Var[2]+Var[4]+Var[6]+sum(Cov)-Cov[0]-Cov[4]-Cov[8],Cov[2]+Cov[3]+Cov[7],Var[2]+Var[4]+Var[6],Var[4]+Cov[2]+Cov[6],Var[2]+Cov[5]+Cov[7],Var[4]+Cov[3]+Cov[5],Var[6]+Cov[6]+Cov[7]],
                     [Var[1]+Var[4]+Var[7]+sum(Cov)-Cov[1]+Cov[2]-Cov[4]-Cov[7],Var[2]+Cov[3]+Cov[8],Var[4]+Cov[2]+Cov[6],Var[2]+Var[4]+Var[7],Cov[2]+Cov[5]+Cov[8],Var[4]+Cov[3]+Cov[5],Var[7]+Cov[6]+Cov[8]],
                     [Var[2]+Var[5]+Var[8]+sum(Cov)-Cov[0]-Cov[3]-Cov[6],Var[8]+Cov[2]+Cov[4],Var[2]+Cov[5]+Cov[7],Cov[2]+Cov[5]+Cov[8],Var[2]+Var[5]+Var[8],Var[5]+Cov[4]+Cov[5],Var[8]+Cov[7]+Cov[8]],
                     [sum(Var[3:6])+2*sum(Cov[3:6]),Var[3]+Cov[3]+Cov[4],Var[4]+Cov[3]+Cov[5],Var[4]+Cov[3]+Cov[5],Var[5]+Cov[4]+Cov[5],sum(Var[3:6])+2*sum(Cov[3:6]),0],
                     [sum(Var[6:9])+2*sum(Cov[6:9]),Var[8]+Cov[7]+Cov[8],Var[6]+Cov[6]+Cov[7],Var[7]+Cov[6]+Cov[8],Var[8]+Cov[7]+Cov[8],0,sum(Var[6:9])+2*sum(Cov[6:9])]
                    ])
    elif cros_type=='BACACBBCA':        
        V = pi*np.array([[sum(Var)+2*sum(Cov),Var[0]+Var[5]+Var[6]+sum(Cov)-Cov[2]-Cov[3]-Cov[8],Var[2]+Var[4]+Var[7]+sum(Cov)-Cov[0]-Cov[4]-Cov[7],Var[1]+Var[4]+Var[7]+sum(Cov)-Cov[1]-Cov[4]-Cov[7],Var[2]+Var[5]+Var[8]+sum(Cov)-Cov[0]-Cov[3]-Cov[6],sum(Var[3:6])+2*sum(Cov[3:6]),sum(Var[6:9])+2*sum(Cov[6:9])],
                         [Var[0]+Var[5]+Var[6]+sum(Cov)-Cov[2]-Cov[3]-Cov[8],Var[0]+Var[5]+Var[6] ,Cov[1]+Cov[5]+Cov[6] ,Cov[0]+Cov[5]+Cov[6] ,Var[5]+Cov[1]+Cov[7] ,Var[5]+Cov[4]+Cov[5] , Var[6]+Cov[6]+Cov[7]],
                         [Var[2]+Var[4]+Var[7]+sum(Cov)-Cov[0]-Cov[4]-Cov[7],Cov[1]+Cov[5]+Cov[6] ,Var[2]+Var[4]+Var[7] ,Var[4]+Var[7]+Cov[2] ,Var[2]+Cov[5]+Cov[8] ,Var[0]+Var[5]+Var[6]+sum(Cov)-Cov[2]-Cov[3]-Cov[8] , Var[7]+Cov[6]+Cov[8]],
                         [Var[1]+Var[4]+Var[7]+sum(Cov)-Cov[1]-Cov[4]-Cov[7],Cov[0]+Cov[5]+Cov[6] ,Var[4]+Var[7]+Cov[2] ,Var[1]+Var[4]+Var[7] ,Cov[2]+Cov[5]+Cov[8] , Var[4]+Cov[3]+Cov[5] , Var[7]+Cov[6]+Cov[8]],
                         [Var[2]+Var[5]+Var[8]+sum(Cov)-Cov[0]-Cov[3]-Cov[6], Var[5]+Cov[1]+Cov[7],Var[2]+Cov[5]+Cov[8] ,Cov[2]+Cov[5]+Cov[8] ,Var[5]+Cov[4]+Cov[5], Var[8]+Cov[7]+Cov[8] ],
                         [sum(Var[3:6])+2*sum(Cov[3:6]), Var[5]+Cov[4]+Cov[5],Var[0]+Var[5]+Var[6]+sum(Cov)-Cov[2]-Cov[3]-Cov[8] ,Var[4]+Cov[3]+Cov[5] ,Var[5]+Cov[4]+Cov[5] , sum(Var[3:6])+2*sum(Cov[3:6]),0],
                         [sum(Var[6:9])+2*sum(Cov[6:9]), Var[6]+Cov[6]+Cov[7] , Var[7]+Cov[6]+Cov[8],Var[7]+Cov[6]+Cov[8] ,Var[8]+Cov[7]+Cov[8] ,0 ,sum(Var[6:9])+2*sum(Cov[6:9])]
                         ])
        
    elif cros_type=='BBAACBCAC':
        V = pi*np.array([[sum(Var)+2*sum(Cov),Var[0]+Var[1]+Var[5]+sum(Cov[0:7])+Cov[0]-Cov[3]+Cov[5],Var[4]+Var[6]+Var[8]+sum(Cov[3:9])-Cov[4]+Cov[7],Var[1]+Var[4]+Var[7]+sum(Cov)-Cov[1]-Cov[4]-Cov[7],Var[2]+Var[5]+Var[8]+sum(Cov)-Cov[0]-Cov[3]+Cov[6],sum(Var[3:6])+2*sum(Cov[3:6]),sum(Var[6:9])+2*sum(Cov[6:9])],
                         [Var[0]+Var[1]+Var[5]+sum(Cov[0:7])+Cov[0]-Cov[3]+Cov[5],Var[0]+Var[1]+2*Cov[0]+Cov[5],Cov[5],Var[1]+Cov[0]+Cov[5],Cov[1]+Cov[2]+Var[5],Var[5]+Cov[4]+Cov[5],0],
                         [Var[4]+Var[6]+Var[8]+sum(Cov[3:9])-Cov[4]+Cov[7],Cov[5],Var[4]+Var[6]+Var[8]+2*Cov[7],Var[4]+Cov[6]+Cov[8],Var[8]+Cov[5]+Cov[7],Var[4]+Cov[3]+Cov[5],Var[6]+Var[8]+sum(Cov[6:9])+Cov[7]],
                         [Var[1]+Var[4]+Var[7]+sum(Cov)-Cov[1]-Cov[4]-Cov[7],Var[1]+Cov[0]+Cov[5],Var[4]+Cov[6]+Cov[8],Var[1]+Var[4]+Var[7],Cov[2]-Cov[5]+Cov[8],Var[4]+Cov[3]+Cov[5],Var[7]+Cov[6]+Cov[8]],
                         [Var[2]+Var[5]+Var[8]+sum(Cov)-Cov[0]-Cov[3]+Cov[6],Cov[1]+Cov[2]+Var[5],Var[8]+Cov[5]+Cov[7],Cov[2]-Cov[5]+Cov[8],Var[2]+Var[5]+Var[8],Var[5]+Cov[7]-Cov[8],Var[8]+Cov[7]+Cov[8]],
                         [sum(Var[3:6])+2*sum(Cov[3:6]),Var[5]+Cov[4]+Cov[5],Var[4]+Cov[3]+Cov[5],Var[7]+Cov[6]+Cov[8],Var[5]+Cov[7]-Cov[8],sum(Var[3:6])+2*sum(Cov[3:6]),0],
                         [sum(Var[6:9])+2*sum(Cov[6:9]),0,Var[6]+Var[8]+sum(Cov[6:9])+Cov[7],Var[7]+Cov[6]+Cov[8],Var[8]+Cov[7]+Cov[8],0,sum(Var[6:9])+2*sum(Cov[6:9])],
                         ])
    return V
    

def Covariance(cros_type,data):
    if len(cros_type)==4:
        Cov=[ data.Yi11.cov(data.Yi12), data.Yi21.cov(data.Yi22)]  
        
    elif len(cros_type)==6:
        Cov=[ data.Yi11.cov(data.Yi12), data.Yi11.cov(data.Yi13),data.Yi12.cov(data.Yi13), 
              data.Yi21.cov(data.Yi22), data.Yi21.cov(data.Yi23),data.Yi22.cov(data.Yi23)]
    elif len(cros_type)==9:
        Cov=[ data.Yi11.cov(data.Yi12), data.Yi11.cov(data.Yi13),data.Yi12.cov(data.Yi13), 
              data.Yi21.cov(data.Yi22), data.Yi21.cov(data.Yi23),data.Yi22.cov(data.Yi23),
              data.Yi31.cov(data.Yi32), data.Yi31.cov(data.Yi33),data.Yi32.cov(data.Yi33)]  
    return Cov

  


sim_time,gamma_param,seq_size=2000,0.1,50
pi=(seq_size)/(seq_size*3)

design_type='ABCBCACAB'
init_params=[0.9, 0.95, 0.95, 0.25, 0.25, 0.15, 0.15]
#tm or estimate
par_tv=[1.0, 1.21, 1.27, 0.32, 0.37, 0.22, 0.28]
true_mean=Mean_Poisson(cros_type=design_type, params_value=par_tv)
I=MatrixI(cros_type=design_type, Pi=pi, tm=true_mean)


np.random.seed(980716)
mle_ind,mle_cor=pd.DataFrame(),pd.DataFrame()
I_ind, V_ind,I_cor, V_cor = 0, 0, 0, 0

for i in range(sim_time):    
    #independent
    df_ind=Data_Generate(cros_type=design_type, mean_true=true_mean, sample_size=seq_size, data_type='ind', cor_par=gamma_param, eta0=1)
    mle_i=MLE(cros_type=design_type, data=df_ind, initial_guess=init_params)
    mle_ind = mle_ind.append(mle_i,ignore_index=True)
    est_mean=Mean_Poisson(cros_type=design_type, params_value=mle_i.to_numpy()[0])
    I_i=MatrixI(cros_type=design_type, Pi=pi, tm=est_mean)
    df_ind_var=df_ind.var(ddof=1)
    df_ind_cov=Covariance(cros_type=design_type,data=df_ind)
    V_i=MatrixV(cros_type=design_type,Pi=pi,Cov=df_ind_cov,Var=df_ind_var)
    I_ind+=I_i
    V_ind+=V_i
    #correlated
    df_cor=Data_Generate(cros_type=design_type, mean_true=true_mean, sample_size=seq_size, data_type='cor', cor_par=gamma_param, eta0=1)
    mle_i=MLE(cros_type=design_type, data=df_cor, initial_guess=init_params)
    mle_cor = mle_cor.append(mle_i,ignore_index=True)
    est_mean=Mean_Poisson(cros_type=design_type, params_value=mle_i.to_numpy()[0])
    I_i=MatrixI(cros_type=design_type, Pi=pi, tm=est_mean)
    df_cor_var=df_cor.var(ddof=1)
    df_cor_cov=Covariance(cros_type=design_type,data=df_cor)
    V_i=MatrixV(cros_type=design_type,Pi=pi,Cov=df_cor_cov,Var=df_cor_var)
    I_cor+=I_i
    V_cor+=V_i
    

I_ind=I_ind/sim_time
V_ind=V_ind/sim_time
I_cor=I_cor/sim_time
V_cor=V_cor/sim_time
np.set_printoptions(suppress=True,precision=5)
print('Independent Data, seq_size = ',seq_size)
print('True Value of parameters\n alpha：%f, eta：%f, gamma：%f, delta：%f' %(1.0, 0.5, 0.2, 0.2))
print('MLE\n',mle_ind.mean())
print('Sample variance of estimates\n',3*seq_size*mle_ind.var(ddof=1))

print('Inverse of matrix I\n',lin.inv(I_ind))
print('S\n',mle_ind.cov(ddof=1))
print('N*S\n',3*seq_size*mle_ind.cov(ddof=1))
print('inv(I)*V*inv(I)\n',lin.inv(I_ind).dot(V_ind).dot(lin.inv(I_ind)))

print('Estimate of matrix I\n',I_ind)
print('Estimate of matrix V\n',V_ind)

print('Correlated Data (alpha =10 ,beta = 0.1), seq_size = ',seq_size)
print('True Value of parameters\n alpha：%f, eta：%f, gamma：%f, delta：%f' %(1.0, 0.5, 0.2, 0.2))#1.0, 1.0, 1.0, 0.5
print('MLE\n',mle_cor.mean())
print('Sample variance of estimates\n',3*seq_size*mle_cor.var(ddof=1))

print('Inverse of matrix I\n',lin.inv(I_cor))
print('S\n',mle_cor.cov(ddof=1))
print('N*S\n',3*seq_size*mle_cor.cov(ddof=1))
print('inv(I)*V*inv(I)\n',lin.inv(I_cor).dot(V_cor).dot(lin.inv(I_cor)))

print('Estimate of matrix I\n',I_cor)
print('Estimate of matrix V\n',V_cor)

import openpyxl
wb = openpyxl.Workbook()
wb.create_sheet("n=100")
my_sheet = wb.active
c1 = my_sheet.cell(row = 1, column = 1)
c1.value = "Independent Data, seq_size = 50"
c2 = my_sheet.cell(row = 2, column = 1)
c2.value='MLE'
MLE_contra = np.array([[1.0, 0.5, 0.2, 0.2], mle_ind.mean().tolist()])
for row in MLE_contra:
    my_sheet.append(row.tolist())

c3 = my_sheet.cell(row=6,column = 1)
c3.value = "inverse I_ind "
for row in lin.inv(I_ind):
    my_sheet.append(row.tolist())
c4 = my_sheet.cell(row=12, column = 1)
c4.value = "N*S"
for row in 2*seq_size*mle_ind.cov(ddof=1):
    my_sheet.append(row.tolist())
c5 = my_sheet.cell(row=18, column = 1)
c5.value = "inv(I_ind)*V*inv(I_ind)"
for row in lin.inv(I_ind).dot(V_ind).dot(lin.inv(I_ind)):
    my_sheet.append(row.tolist())
wb.save('ABBA.xlsx')
